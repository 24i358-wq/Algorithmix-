"""
make_sample_input.py
=====================
This is a HELPER script only - it is NOT part of the main analyzer.

The main tool (risk_analyzer_gui.py) expects an Excel file as INPUT.
Since we don't have a real company spreadsheet, this script creates a
sample one called "dependency_data.xlsx" so you have something to open
in the GUI and test with.

Run once:
    python3 make_sample_input.py

It creates: dependency_data.xlsx  (open this from inside the GUI)
"""

import random
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

random.seed(42)
TODAY = date(2026, 7, 11)

LIBRARY_NAMES = [
    "log4j-core", "jackson-databind", "spring-core", "guava", "commons-lang3",
    "requests", "flask", "django", "numpy", "pandas", "lodash", "express",
    "react", "axios", "left-pad", "chalk", "openssl-wrapper", "protobuf",
]
SAFE_LICENSES = ["MIT", "Apache-2.0", "BSD-3-Clause"]
RISKY_LICENSES = ["GPL-3.0", "AGPL-3.0", "Unlicensed"]
ALL_LICENSES = SAFE_LICENSES + RISKY_LICENSES

# a few "known broken" library+version pairs so the sample data has real hits
KNOWN_BROKEN_VERSIONS = [
    ("log4j-core", "2.14.1"), ("jackson-databind", "2.9.8"), ("spring-core", "5.2.0"),
    ("commons-lang3", "3.8"), ("flask", "0.12"), ("left-pad", "1.0.0"), ("openssl-wrapper", "1.1.0"),
]

APP_NAMES = ["PaymentGateway", "CoreLedger", "MobileBankingApp", "RiskEngine",
             "TradeSettlement", "CustomerDataHub", "IAMPortal", "OnboardingPortal"]

rows = []
for i, app_name in enumerate(APP_NAMES, start=1):
    app_id = f"APP-{i:03d}"
    distributed = "No" if app_name == "IAMPortal" else "Yes"

    for _ in range(15):
        library = random.choice(LIBRARY_NAMES)
        version = f"{random.randint(1,5)}.{random.randint(0,9)}.{random.randint(0,20)}"
        license_name = random.choice(ALL_LICENSES)
        days_ago = random.randint(0, 4 * 365)
        last_updated = (TODAY - timedelta(days=days_ago)).isoformat()
        rows.append([app_id, app_name, library, version, license_name, "direct", last_updated, distributed])

    for _ in range(5):
        library = random.choice(LIBRARY_NAMES)
        version = f"{random.randint(1,5)}.{random.randint(0,9)}.{random.randint(0,20)}"
        license_name = random.choice(ALL_LICENSES)
        days_ago = random.randint(0, 4 * 365)
        last_updated = (TODAY - timedelta(days=days_ago)).isoformat()
        rows.append([app_id, app_name, library, version, license_name, "transitive", last_updated, distributed])

# plant some known-broken versions so the demo has real "Critical" hits
for row in random.sample(rows, 25):
    library, version = random.choice(KNOWN_BROKEN_VERSIONS)
    row[2], row[3] = library, version

wb = Workbook()
ws = wb.active
ws.title = "Dependencies"
headers = ["App ID", "App Name", "Library", "Version", "License",
           "Dependency Type", "Last Updated", "Distributed"]
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1A2A44", end_color="1A2A44", fill_type="solid")

for row in rows:
    ws.append(row)

for col, width in zip("ABCDEFGH", [10, 18, 18, 10, 12, 15, 14, 12]):
    ws.column_dimensions[col].width = width

wb.save("dependency_data.xlsx")
print(f"Created dependency_data.xlsx with {len(rows)} dependency rows across {len(APP_NAMES)} apps.")
